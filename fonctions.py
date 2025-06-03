import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import re
from datetime import datetime


def df_journaux(jSOCIETE):
    df_journaux = pd.read_excel("datas\mapping code journaux.xlsx", sheet_name=jSOCIETE)
    return df_journaux


def ajouter_type_piece5(df, societe):
    # Générer la CléUnique
    df['CléUnique'] = np.where(
        df['Code journal'] == 'AN',
        'AN',
        np.where(
            df['Code journal'] == 'RAN',
            'RAN',
            df['Code journal'] + df['N° pièce'].astype(str)
        )
    )

    df['type_piece'] = ""

    # Définir les règles pour les préfixes de journal et de comptes
    rules = [
        (('BAN', 'BQ', 'NS', 'SG'), ('52',), "BAN group"),
        (('CAI', 'CS'), ('572', '571'), "CAI group"),
        (('PMN', 'PNM'), ('554',), "PMN group"),
        (('FINAFF',), ('561',), "FINAFF group"),
    ]

    # Appliquer les règles par groupe
    for piece, group in df.groupby("CléUnique"):
        type_piece = ""

        for index, row in group.iterrows():
            code_journal = str(row['Code journal'])
            compte_gen = str(row['N° compte général'])
            debit = row['Débit']
            credit = row['Crédit']

            for prefixes_journal, prefixes_compte, _ in rules:
                if any(code_journal.startswith(prefix) for prefix in prefixes_journal):
                    if any(compte_gen.startswith(p) for p in prefixes_compte):
                        if pd.isna(debit) or debit == "":
                            type_piece = "DECAI"
                        elif pd.isna(credit) or credit == "":
                            type_piece = "ENCAI"
                        else:
                            type_piece = "XXXXXXXXXXX"
                            print("NOTHING !!!")
                        break
            if type_piece:
                break  # Dès qu'on trouve une correspondance, inutile de continuer

        df.loc[group.index, 'type_piece'] = type_piece

    # Règles directes de correspondance
    mapping = {
        'AA': 'FAFOU', 'AC': 'FAFOU', 'ABN': 'ODDIV', 'AB': 'ODDIV',
        'ACH': 'FAFOU', 'VE': 'FACLI', 'VTE': 'FACLI', 
        'AFFAC': 'ODDIV',
        'IMMO': 'FAFOU',
        'OD': 'ODDIV', 'ND': 'ODDIV',
        'STKCDG': 'STOCK',
        'PAI': 'ODPAI', 'PAIAP': 'ODPAI', 'PA': 'ODPAI',
        'AN': 'RANX1', 'RAN': 'RANX1',
        'CHQ': 'ENCAI'
    }

    for code_journal, type_val in mapping.items():
        df.loc[df['Code journal'] == code_journal, 'type_piece'] = type_val

    # Pourcentage de lignes sans type
    x = df['type_piece'].eq('').sum()
    y = len(df)
    pourcentage = round((x / y) * 100, 3)
    print(f"Il y a {pourcentage}% de lignes sans type de pièce")
    print(f"Les pieces sans type de piece sont : {list(df[df['type_piece']=='']['CléUnique'])}")
    df.loc[df['type_piece'] == '', 'type_piece'] = 'ODDIV'
    if societe=="LCR":
        df.loc[df['type_piece']=="FAFOU", 'type_piece'] = "FAF01"
        df.loc[df['type_piece']=="ODDIV", 'type_piece'] = "ODD01"
        df.loc[df['type_piece']=="FACLI", 'type_piece'] = "FAC01"
    elif societe=="MIN":
        df.loc[df['type_piece']=="FAFOU", 'type_piece'] = "FAF02"
        df.loc[df['type_piece']=="ODDIV", 'type_piece'] = "ODD02"
        df.loc[df['type_piece']=="FACLI", 'type_piece'] = "FAC02"
    elif societe=="LOG":
        df.loc[df['type_piece']=="FAFOU", 'type_piece'] = "FAF03"
        df.loc[df['type_piece']=="ODDIV", 'type_piece'] = "ODD03"
        df.loc[df['type_piece']=="FACLI", 'type_piece'] = "FAC03"
    elif societe=="SCI":
        df.loc[df['type_piece']=="FAFOU", 'type_piece'] = "FAF04"
        df.loc[df['type_piece']=="ODDIV", 'type_piece'] = "ODD04"
        df.loc[df['type_piece']=="FACLI", 'type_piece'] = "FAC04"
    elif societe=="GC":
        df.loc[df['type_piece']=="FAFOU", 'type_piece'] = "FAF05"
        df.loc[df['type_piece']=="ODDIV", 'type_piece'] = "ODD05"
        df.loc[df['type_piece']=="FACLI", 'type_piece'] = "FAC05"
    
    return df


def split_section_analytique(df):
    df = df.copy()  # Évite de modifier le dataframe original
    df['Libellé écriture'] = df['Libellé écriture'].str.replace(';', ':', regex=False)
    

    df["Axe1"] = df["section analytique"].apply(lambda x: str(x)[:3] if pd.notna(x) else "")
    df["Axe2"] = df["section analytique"].astype(str).str[3:6]  # 3 suivants
    df["Axe3"] = df["section analytique"].apply(lambda x: "00" + str(x)[-2:] if pd.notna(x) else "")  # "00" + 2 derniers
    df["Axe4"] = ""

    mapping = {
        'CCF': 'CCFA',
        'CEU': 'CEUR',
        'EUR': 'CEUR',
        'CMA': 'CMAR',
        'CNE': 'CNES',
        'CSI': 'CSIT',
        'CSO': 'CSOL',
        'SGM': 'DSMG',
        'COL': 'FCO',
        'FAL': '???'
    }

    df['Axe2'] = df['Axe2'].replace(mapping)
    
    return df


def completer_date_echeance(df_x3):
    for piece, group in df_x3.groupby("N° pièce"):
        # Récupérer la dernière date non vide du groupe
        date_echeance = group['Date échéance'].dropna().iloc[-1] if group['Date échéance'].notna().any() else None

        # Remplir les valeurs NaN avec la date trouvée
        df_x3.loc[df_x3["N° pièce"] == piece, "Date échéance"] = date_echeance

    return df_x3


def merge_custom(df_left, df_right, key):
    # Étape 1 : Supprimer les doublons sur la clé dans df_right
    df_right = df_right.drop_duplicates(subset=key, keep='first')

    # Étape 2 : Créer un dictionnaire des valeurs de droite (sans la clé)
    right_dict = df_right.set_index(key).to_dict(orient='index')

    # Étape 3 : Créer des colonnes vides à remplir
    new_columns = {col: [] for col in df_right.columns if col != key}

    # Étape 4 : Parcourir les lignes du DataFrame gauche
    for _, row in df_left.iterrows():
        join_key = row.get(key)
        match = right_dict.get(join_key)

        if match:
            for col in new_columns:
                new_columns[col].append(match.get(col))
        else:
            for col in new_columns:
                new_columns[col].append(pd.NA)

    # Étape 5 : Ajouter les nouvelles colonnes au DataFrame gauche
    for col, values in new_columns.items():
        df_left[col] = values

    return df_left


def merge_Tiers2(df, dftiers):
    #df["N° compte tiers"] = df["N° compte tiers"].astype(str).str.strip()
    #dftiers["N° compte tiers"] = dftiers["N° compte tiers"].astype(str).str.strip()
    #df["N° compte tiers"] = df["N° compte tiers"].astype(str)
    dftiers["N° compte tiers"] = dftiers["N° compte tiers"].astype(str)
    #print(dftiers["N° compte tiers"].value_counts())
    #print(df["N° compte tiers"].isna().sum())
    #print("Les colonnes de Dftiers sont : ", dftiers.columns)
    dftiers = dftiers[['N° compte tiers', 'NouveauCode']]

    # Créer un dictionnaire de mapping
    #mapping = dict(zip(dftiers["N° compte tiers"], dftiers["NouveauCode"]))
    #print('Mapping SUCCESS')
    mapping = dftiers.set_index("N° compte tiers")["NouveauCode"].to_dict()
    df["Tier"] = df["N° compte tiers"].map(mapping)

    #Ajouter une colonne 'Tier' dans df en remplaçant selon le mapping
    #df["Tier"] = df["N° compte tiers"].replace(mapping)
    #df["Tier"] = df["N° compte tiers"].map(mapping)
    #print('Fonction merged tiers OK')
    #print(dftiers["NouveauCode"])

    #Extraire les lignes non affectées (Tier == NaN)
    df_non_affectes = df[(df["N° compte tiers"].notna() | df["N° compte tiers"]=="nan") & (df["Tier"].isna())]
    
    return df, df_non_affectes


def merge_Tiers3(df, dftiers):
    # Nettoyage des colonnes
    dftiers["N° compte tiers"] = dftiers["N° compte tiers"].astype(str).str.strip()
    df["N° compte tiers"] = df["N° compte tiers"].astype(str).str.strip()
    dftiers = dftiers.drop_duplicates(subset="N° compte tiers", keep="first")
    # Ne garder que les colonnes utiles
    dftiers = dftiers[['N° compte tiers', 'NouveauCode']]
    if "Tiers" in df.columns:
        df = df.drop(columns=["Tiers"])

    # Utiliser notre fonction custom merge
    print(df.columns)
    print('*****************************************************************')
    print(dftiers.columns)
    df = merge_custom(df, dftiers, key="N° compte tiers")
    print('*****************************************************************')
    print(df.columns)

    

    # Renommer la colonne obtenue
    #df = df.rename(columns={"NouveauCode": "Tier"})

    # Extraire les lignes non affectées
    df_non_affectes = df
    #df[(df["N° compte tiers"].notna() | (df["N° compte tiers"] == "nan")) & (df["Tier"].isna())]

    return df, df_non_affectes


def merge_journaux(df, dfjournaux):
  dfjournaux.rename(columns={"NouveauCode": "BonCode_Journal"}, inplace=True)
  df_merged = df.merge(dfjournaux, on="Code journal", how="left")
  return df_merged


def merge_IFRS(df, dfIFRS):
  dfIFRS.rename(columns={"Compte local": "N° compte général"}, inplace=True)
  df_merged = df.merge(dfIFRS, on="N° compte général", how="left")
  return df_merged


def retriev_code_PNL_from_df(df, pnl):
  df["Axe1"] = df["Axe1"].fillna("").apply(
        lambda x: str(int(x)) if isinstance(x, (int, float)) and not pd.isna(x) and x == int(x) else str(x)
    )
  
  df["Critère 2"] = df.apply(
        lambda row: f"{row['Axe1']} {row['Axe2']} {row['N° compte général']}"
        if pd.notna(row["section analytique"]) and row["section analytique"] != "" else "", axis=1
    )
  pnl["Critère 2"] = pnl["Critère 2"].astype(str)
  df_PNL = df.merge(pnl[["Critère 2", "Mapping crit 1", "Code1"]], on="Critère 2", how="left")
  df_PNL.rename(columns={"Code1": "Code_PNL"}, inplace=True)
  return df_PNL


def renommer_colonnes(data):
  data.rename(columns={"BonCode_Journal": "Code_Journal"}, inplace=True)
  data.rename(columns={"Tier": "Tiers"}, inplace=True)
  data.rename(columns={"Rubriques OPERA": "Code_IFRS"}, inplace=True)
  data.rename(columns={"Mapping crit 1": "Designation_PNL"}, inplace=True)
  return data


def preparation2(df, dftiers, dfjournaux, dfIFRS, pnl, societe):
    # Étape 1 : Définir les noms de colonnes attendus
    colonnes_attendues = [
        'Code journal', 'Date', 'N° pièce', 'N° facture', 'Référence',
        'N° compte général', 'N° compte tiers', 'Libellé écriture',
        'Date échéance', 'Débit', 'Crédit', 'section analytique'
    ]

    # Étape 2 : Vérification du nombre de colonnes
    if df.shape[1] != len(colonnes_attendues):
        raise ValueError(
            f"Le DataFrame a {df.shape[1]} colonnes, mais {len(colonnes_attendues)} sont attendues.\n"
            f"Colonnes actuelles : {list(df.columns)}\n"
            f"Colonnes attendues : {colonnes_attendues}"
        )

    # Étape 3 : Vérification si les colonnes ont bien les bons noms
    colonnes_incorrectes = [col for col in df.columns if col not in colonnes_attendues]

    if colonnes_incorrectes:
        print("⚠️ Colonnes mal nommées ou incorrectes détectées. Tentative de correction par position...")
        print(f"Colonnes actuelles : {list(df.columns)}")
        print(f"Colonnes attendues : {colonnes_attendues}")
        df.columns = colonnes_attendues  # Remplacement basé sur la position

    # Étape 4 : Lancement du traitement
    df1 = ajouter_type_piece5(df, societe)
    df11 = completer_date_echeance(df1)
    df2 = split_section_analytique(df11)

    print("Nombre de lignes avant merge_Tiers:", len(df2))

    df3, _ = merge_Tiers2(df2, dftiers)
    df4 = merge_journaux(df3, dfjournaux)
    df5 = df4.copy()
    df6 = renommer_colonnes(df5)

    # Clé unique
    df6['CléUnique'] = np.where(
        df6['Code journal'] == 'AN', 'AN',
        np.where(df6['Code journal'] == 'RAN', 'RAN',
                 df6['Code journal'] + df6['N° pièce'].astype(str))
    )

    return df6


def convertir_sage100_en_x31(df_sage100, societe='LCR'):
  df_sage100 = df_sage100.loc[:, ~df_sage100.columns.duplicated()]
  df_Col = [str(i) for i in range(1, 15)]
  df = pd.DataFrame(columns=df_Col)
  #df_sage100['CléUnique'] = df_sage100['Code journal'] + df_sage100['N° pièce'].astype(str)
  df_sage100['CléUnique'] = np.where(
    df_sage100['Code journal'] == 'AN',
    'AN',
    np.where(
        df_sage100['Code journal'] == 'RAN',
        'RAN',
        df_sage100['Code journal'] + df_sage100['N° pièce'].astype(str)
    )
  )
  df_sage100['CléUnique2'] = df_sage100['Code_Journal'] + df_sage100['N° pièce'].astype(str)
  #LIGNE D'ENTETE
  for cle, group in df_sage100.groupby("CléUnique"):
        df_E = []
        df_D = []

        date_comptable = pd.to_datetime(group["Date"].iloc[0], errors='coerce')
        date_echeance = pd.to_datetime(group["Date échéance"].iloc[0], errors='coerce')

        type_de_piece = group["type_piece"].iloc[0]
        numero_de_piece = "" #LAISSER VIDE
        match societe :
          case 'LCR':
            site = "LCR00"
          case 'MIN':
            site = "MIN31"
          case 'LOG':
            site = "LOG01"
          case 'SCI':
            site = "LAV21"
          case 'GC':
            site = "GCS00"
        journal = group["Code_Journal"].iloc[0]
        
        date_comptable = date_comptable.strftime('%y%m%d') if pd.notna(date_comptable) else ""
        categorie = 2
        date_echeance = date_echeance.strftime('%y%m%d') if pd.notna(date_echeance) else ""
        numerofacture = group["N° facture"].iloc[0]
        reference1 = numerofacture
        devise = "XOF"
        transaction = "STDCO"
        reference2 = group["Référence"].iloc[0]

        ligne_entete = ["G", type_de_piece, numero_de_piece, site, journal, date_comptable, categorie,
                        date_echeance, reference1, reference2, devise, transaction,  "", cle]
        df_E.append(ligne_entete)
        #print("entete TERMINé")
        #LIGNE DE DETAIL
        numero_ligne = 0
        num_ordre = 0
        for index, row in group.iterrows():
            type_ref = 1
            match societe :
                case 'LCR':
                  site = "LCR00"
                case 'MIN':
                  site = "MIN31"
                case 'LOG':
                  site = "LOG01"
                case 'SCI':
                  site = "LAV21"
                case 'GC':
                  site = "GCS00"
            collectif = "" #LAISSER VIDE
            compte_general = row["N° compte général"]
            #code_IFRS = row["Code_IFRS"]
            val_tiers = row.get('Tiers', "")
            #print(val_tiers)
            compte_tiers = str(val_tiers) if pd.notna(val_tiers) and str(val_tiers).strip() != "" else ""
            #compte_tiers = row['Tiers'] if pd.notna(row['Tiers']) else ""
            libelle = row["Libellé écriture"]
            sens = -1 if pd.notna(row["Crédit"]) else 1
            montant = row["Crédit"] if pd.notna(row["Crédit"]) else row["Débit"]

            numero_ligne += 1
            ligne_detail = ["D", numero_ligne, type_ref, site, collectif,
                            compte_general, compte_tiers, libelle, sens, montant, devise, "", "", cle]
            df_D.append(ligne_detail)

            #ANALYTIQUES
            axe1 = row.get("Axe1", "")
            if pd.notna(axe1) and str(axe1).strip() != "":
              #numero_ligne += 1
              ligne_detail_ana = ["D", numero_ligne, 2, site, collectif,
                            compte_general, compte_tiers, libelle, sens, montant, devise, "", "", cle]
              df_D.append(ligne_detail_ana)
              num_ordre += 1
              ligne_Ana = ["A", num_ordre, "1NC", row["Axe1"], "2BU", row["Axe2"],
                            "3AG", row["Axe3"], "4RE", row["Axe4"], montant, "", "", cle]
              df_D.append(ligne_Ana)
              #df_A.append(ligne_Ana)
              #df_A.append(ligne_detail_ana)
            #numero_ligne += 1
            #if code_IFRS != "" and code_IFRS != "nan":
              #ligne_detail_IFRS = ["D", numero_ligne, 3, site, collectif, code_IFRS,
                             # "",  libelle, sens, montant, devise, "", "", cle]
              #df_D.append(ligne_detail_IFRS)

        df_x3 = df_E + df_D #+ df_A
        df = pd.concat([df, pd.DataFrame(df_x3, columns=df_Col)], ignore_index=True)
  df.rename(columns={"14": "N° pièce"}, inplace=True)
  return df


def qualifier_et_controler3(df_sage100, df_sageX3):
    import pandas as pd
    import numpy as np

    # Initialisation du DataFrame résultat
    colonnes = ["N° pièce", "clé2", "nb de Tiers sage 100", "nb de Tiers sage X3", "liste tiers sage 100", "liste tiers sage X3", "validation_tiers",
                "nb de CG sage 100", "nb de CG sage X3", "liste CG sage 100", "liste CG sage X3", "validation_cg",
                "Total_débit_s100", "Total_crédit_s100", "Solde_s100", "équilibre100", "Total_débit_sX3", "Total_crédit_sX3", "Solde_sX3", "équilibreX3", "validation_solde"]
    df_qualite = pd.DataFrame(columns=colonnes)
    liste_tous_les_tiers_non_trouves = []

    df_sage100['CléUnique'] = np.where(
        df_sage100['Code journal'] == 'AN', 'AN',
        np.where(df_sage100['Code journal'] == 'RAN', 'RAN',
                 df_sage100['Code journal'] + df_sage100['N° pièce'].astype(str))
    )

    pieces_sage100 = df_sage100.groupby("CléUnique")
    pieces_sageX3 = df_sageX3.groupby("N° pièce")

    for cle in set(df_sage100["CléUnique"].unique()):
        somme_debit100 = somme_credit100 = solde_sage100 = 0
        somme_debitX3 = somme_creditX3 = solde_sageX3 = 0

        list_tiers_sage100 = set()
        cle2 = df_sage100["CléUnique"].iloc[0]
        list_cg_sage100 = set()

        if cle in pieces_sage100.groups:
            group_s100 = pieces_sage100.get_group(cle)
            list_tiers_sage1002 = set(group_s100.loc[group_s100["N° compte tiers"].notna() & (group_s100["N° compte tiers"] != ""), "N° compte tiers"])
            list_cg_sage100 = set(group_s100.loc[group_s100["N° compte général"].notna() & (group_s100["N° compte général"] != ""), "N° compte général"])
            somme_debit100 = group_s100["Débit"].sum()
            somme_credit100 = group_s100["Crédit"].sum()
            solde_sage100 = somme_debit100 - somme_credit100

        list_tiers_sageX3 = set()
        list_cg_sageX3 = set()

        if cle in pieces_sageX3.groups:
            group_sX3 = pieces_sageX3.get_group(cle)
            list_tiers_sageX3 = set(group_sX3.loc[(group_sX3["7"] != "") & (group_sX3["1"] == "D") & (group_sX3["3"] != 3), "7"])
            list_cg_sageX3 = set(group_sX3.loc[(group_sX3["6"] != "") & (group_sX3["1"] == "D") & (group_sX3["3"] != 3), "6"])
            somme_debitX3 = group_sX3.loc[(group_sX3["1"] == "D") & (group_sX3["3"] == 1) & (group_sX3["9"] == -1), "10"].sum()
            somme_creditX3 = group_sX3.loc[(group_sX3["1"] == "D") & (group_sX3["3"] == 1) & (group_sX3["9"] == 1), "10"].sum()
            solde_sageX3 = somme_debitX3 - somme_creditX3

        validation_tiers = len(list_tiers_sage1002) == len(list_tiers_sageX3)
        tiers_non_trouves = list_tiers_sage1002 - list_tiers_sageX3
        if tiers_non_trouves:
            liste_tous_les_tiers_non_trouves.extend(tiers_non_trouves)
            #print(f"Tiers non trouvés pour la pièce {cle} : {tiers_non_trouves}")
        validation_cg = list_cg_sage100 == list_cg_sageX3
        equilibre100 = "pièce équilibrée" if solde_sage100 == 0 else "pièce non équilibrée"
        equilibreX3 = "pièce équilibrée" if solde_sageX3 == 0 else "pièce non équilibrée"
        validation_solde = (solde_sage100 == solde_sageX3) & (equilibre100 == equilibreX3 == "pièce équilibrée")

        list_tiers_sage100_str = ", ".join(sorted(map(str, list_tiers_sage1002)))
        list_tiers_sageX3_str = ", ".join(sorted(map(str, list_tiers_sageX3)))
        list_cg_sage100_str = ", ".join(sorted(map(str, list_cg_sage100)))
        list_cg_sageX3_str = ", ".join(sorted(map(str, list_cg_sageX3)))

        df_qualite.loc[len(df_qualite)] = [
            cle, cle2, len(list_tiers_sage1002), len(list_tiers_sageX3),
            list_tiers_sage100_str, list_tiers_sageX3_str, validation_tiers,
            len(list_cg_sage100), len(list_cg_sageX3),
            list_cg_sage100_str, list_cg_sageX3_str, validation_cg,
            somme_debit100, somme_credit100, solde_sage100, equilibre100,
            somme_debitX3, somme_creditX3, solde_sageX3, equilibreX3,
            validation_solde
        ]

    def separatrice(df, liste):
        df_filtered = df[~df['N° pièce'].isin(liste)]
        df_excluded = df[df['N° pièce'].isin(liste)]
        return df_excluded, df_filtered

    # Pièces avec problème de solde
    listt_solde = list(df_qualite[df_qualite["validation_solde"] == False]["N° pièce"].unique())
    df_excluded, df_filtered = separatrice(df_sageX3, listt_solde)

    # Pièces avec problème de tiers
    listt_tiers = list(df_qualite[df_qualite["validation_tiers"] == False]["N° pièce"].unique())
    df_tiers_excluded, df_tiers_filtered = separatrice(df_filtered, listt_tiers)

    # Résumé
    print("La qualification des données a été effectuée avec SUCCÈS !!!")
    print("Pourcentage de pièces déséquilibrées : ", round((len(df_qualite[~df_qualite["validation_solde"]]) / len(df_qualite)) * 100, 2), "%")
    print("Pièces déséquilibrées : ", listt_solde)
    print("Pourcentage de pièces avec des tiers manquants : ", len(df_qualite[~df_qualite["validation_tiers"]]), "sur ", len(df_qualite) ,  round((len(df_qualite[~df_qualite["validation_tiers"]]) / len(df_qualite)) * 100, 2), "%")
    print("Taux de tiers trouvés (en volume) : ", df_qualite['nb de Tiers sage X3'].sum(), "sur ", df_qualite['nb de Tiers sage 100'].sum() , round((df_qualite['nb de Tiers sage X3'].sum() / df_qualite['nb de Tiers sage 100'].sum()) * 100, 2), "%")
    print("\nListe unique de tous les tiers Sage 100 sans correspondance Sage X3 :")
    print(sorted(set(liste_tous_les_tiers_non_trouves)))

    return df_qualite, df_excluded, df_filtered, df_tiers_excluded, df_tiers_filtered


def nettoyer_dataframe(df):
    df = df.copy()
    df.columns = df.iloc[0]            # Remplacer les noms de colonnes par la première ligne
    df = df[1:]                         # Supprimer cette première ligne devenue inutile
    df = df.iloc[:, :-1]               # Supprimer la dernière colonne
    df = df.reset_index(drop=True)     # Réindexer proprement
    return df


def transformer_ecritures_stock(df):
    # Étape 1 : Renommer les colonnes
    df = df.copy()
    df.columns = [
        'Code journal', 'Date', 'N° pièce', 'N° facture', 'Référence',
        'N° compte général', 'N° compte tiers', 'Libellé écriture', 'Débit',
        'Crédit', 'type_de_ligne', 'section analytique', 'Column13'
    ]

    # Étape 2 : Ajouter colonne 'Piece_unique'
    df['Piece_unique'] = df['Code journal'].astype(str) + df['N° pièce'].astype(str)

    # Résultat final
    pieces_equilibrees = []

    for piece_id, group in df.groupby('Piece_unique'):
        group = group.copy()

        # Ligne G avec compte tiers
        ligne_G = group[
            (group['type_de_ligne'] == 'G') &
            (group['N° compte tiers'].notna()) &
            (group['N° compte tiers'] != '')
        ]

        # Ligne G2 sans compte tiers
        ligne_G2 = group[
            (group['type_de_ligne'] == 'G') &
            ((group['N° compte tiers'].isna()) | (group['N° compte tiers'] == ''))
        ]

        if ligne_G.empty or ligne_G2.empty:
            print(f"Pièce {piece_id} ignorée (ligne G ou G2 manquante)")
            continue

        ligne_G = ligne_G.iloc[0].copy()
        ligne_G2 = ligne_G2.iloc[0].copy()

        # Ligne A avec section analytique valide
        ligne_A = group[
            (group['type_de_ligne'] == 'A') &
            (group['section analytique'].notna()) &
            (group['section analytique'] != '')
        ]

        # Sinon motif spécial
        if ligne_A.empty:
            ligne_A = group[
                (group['type_de_ligne'] == 'A') &
                (group['section analytique'].astype(str).str.match(r'^\d{3}[A-Z]+\d+$'))
            ]
        if ligne_A.empty:
            print(f"Pièce {piece_id} ignorée (aucune ligne A valide)")
            continue

        # Nouvelle ligne A
        nouvelle_ligne = ligne_G.copy()
        nouvelle_ligne['type_de_ligne'] = 'A'
        nouvelle_ligne['section analytique'] = ligne_A.iloc[0]['section analytique']

        # Convertir montants
        for ligne in [ligne_G, ligne_G2, nouvelle_ligne]:
            ligne['Débit'] = pd.to_numeric(ligne['Débit'], errors='coerce') or 0
            ligne['Crédit'] = pd.to_numeric(ligne['Crédit'], errors='coerce') or 0

        # Concaténer : G, A, G2
        group_final = pd.DataFrame([ligne_G, nouvelle_ligne, ligne_G2])

        # Vérifier équilibre
        g_lines = group_final[group_final['type_de_ligne'] == 'G']
        debit_total = g_lines['Débit'].sum()
        credit_total = g_lines['Crédit'].sum()

        if round(debit_total, 2) == round(credit_total, 2):
            print(f"Pièce {piece_id} équilibrée")
            pieces_equilibrees.append(group_final)
        else:
            # Correction : ajuster Crédit de G2
            ligne_G2['Crédit'] = ligne_G['Débit']
            group_final.iloc[2] = ligne_G2

            # Recalcul après correction
            g_lines_corr = group_final[group_final['type_de_ligne'] == 'G']
            debit_total_corr = g_lines_corr['Débit'].sum()
            credit_total_corr = g_lines_corr['Crédit'].sum()

            if round(debit_total_corr, 2) == round(credit_total_corr, 2):
                print(f"Pièce {piece_id} équilibrée après correction")
                pieces_equilibrees.append(group_final)
            else:
                print(f"Pièce {piece_id} non équilibrée même après correction")

    # Résultat unique
    resultat = pd.concat(pieces_equilibrees, ignore_index=True) if pieces_equilibrees else pd.DataFrame(columns=df.columns)
    return resultat


def mettre_en_gras_les_G(fichier_excel, feuilles=None):
    wb = load_workbook(fichier_excel)

    if feuilles is None:
        feuilles = wb.sheetnames  # Toutes les feuilles

    for feuille in feuilles:
        ws = wb[feuille]
        for row in ws.iter_rows():
            # Vérifie si au moins une cellule de la ligne contient "G"
            if any(cell.value == "G" for cell in row):
                # Applique le gras à toute la ligne
                for cell in row:
                    cell.font = Font(bold=True)

    wb.save(fichier_excel)


def pipeline(chemin_df, dftiers, dfjournaux, dfIFRS, pnl, societe, mois, version, DS="datas/Enregistrements"):
    # Lire le fichier d'entrée
    df = pd.read_excel(chemin_df, sheet_name=societe)

    # Traitement
    data = preparation2(df, dftiers, dfjournaux, dfIFRS, pnl, societe)  # à retourner
    dataAN = data[(data["CléUnique"] == "AN") | (data["CléUnique"] == "RAN")]  # à retourner
    data = data[(data["CléUnique"] != "AN") & (data["CléUnique"] != "RAN")]

    dfX3 = convertir_sage100_en_x31(data, societe)  # à retourner
    df_qualite, df_excluded, df_filtered, df_tiers_excluded, df_tiers_filtered = qualifier_et_controler3(data, dfX3)  # à retourner

    # Nettoyage
    dfX3 = nettoyer_dataframe(dfX3)
    df_filtered = nettoyer_dataframe(df_filtered)
    df_tiers_filtered = nettoyer_dataframe(df_tiers_filtered)
    #df_excluded = nettoyer_dataframe(df_excluded)

    # Préparer le chemin
    dossier_sortie = DS
    os.makedirs(dossier_sortie, exist_ok=True)
    annee_actuelle = datetime.now().year
    fichier_sortie = os.path.join(dossier_sortie, f"Ecritures_X3_{mois}{annee_actuelle}_V{version}.xlsx")

    # Écriture conditionnelle
    if os.path.exists(fichier_sortie):
        # Ajouter à un fichier existant
        with pd.ExcelWriter(fichier_sortie, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_filtered.to_excel(writer, sheet_name=societe, index=False)
            df_excluded.to_excel(writer, sheet_name=f"{societe}_Excluded", index=False)
    else:
        # Créer un nouveau fichier
        with pd.ExcelWriter(fichier_sortie, engine='openpyxl') as writer:
            df_filtered.to_excel(writer, sheet_name=societe, index=False)
            df_excluded.to_excel(writer, sheet_name=f"{societe}_Excluded", index=False)

    mettre_en_gras_les_G(fichier_sortie, feuilles=[societe, f"{societe}_Excluded"])

    return data, dataAN, dfX3, df_filtered, df_tiers_filtered, df_excluded, df_tiers_filtered, df_qualite, fichier_sortie


def extraire_infos_depuis_nom(fichier):
    fichier = fichier.lower()
    mois_mapping = {
        "janvier": "janvier", "février": "fevrier", "fevrier":"fevrier", "mars": "mars", "avril": "avril",
        "mai": "mai", "juin": "juin", "juillet": "juillet", "août": "aout", "aout":"aout",
        "septembre": "septembre", "octobre": "octobre", "novembre": "novembre", "décembre": "decembre"
    }

    societe = None
    for s in ['LCR', 'MIN', 'LOG', 'SCI', 'GC']:
        if s.lower() in fichier:
            societe = s
            break

    mois = None
    for mot in mois_mapping:

        if mot in fichier:
            mois = mois_mapping[mot]
            break

    if not mois:
        mois = datetime.now().strftime("%B").lower()

    return societe, mois


def charger_journal_pour_societe(societe):
    jLCR = "mapping_LCR"
    jMIN = "mapping_MINO"
    jLOG = "mapping_LOG"
    jSCI = "mapping_LAVION"
    jGC = "mapping_GRP"
    mapping = {
        "LCR": jLCR,
        "MIN": jMIN,
        "LOG": jLOG,
        "SCI": jSCI,
        "GC": jGC
    }
    if societe in mapping:
        return df_journaux(mapping[societe])
    else:
        raise ValueError(f"Société inconnue ou mapping journal non trouvé pour {societe}")






